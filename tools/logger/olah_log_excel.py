#!/usr/bin/env python3
# olah_log_excel.py — konversi SEMUA file log → satu Excel, tiap log jadi tabel lengkap
#   (seluruh baris anomali, BUKAN rata-rata). Untuk bagian "E. Olah Data Otomatis".
import csv, re, os, sys, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RE_ANOM   = re.compile(r"\[SEC\] >>> ANOMALI @t=(\d+)ms \| slave=(\d+) \| jenis=(\w+)(?:\s*\|\s*detail=(.*))?")
RE_COMMIT = re.compile(r"\[BC\] anomaly committed @t=(\d+)ms \| slave=(\d+)")

# daftar log: (label, nama_file, slave_target, jenis_diharapkan)
LOGS = [
    ("A",   "log_A_new.txt", 2, "ROGUE_ID"),
    ("B",   "log_B.txt",     2, "VALUE_RANGE"),
    ("C",   "log_C.txt",     1, "DEVICE_LOST"),
    ("SUB", "log_SUB2.txt",  2, "IDENTITY"),
    ("D",   "log_SUB.txt",   2, "NONE"),
    ("E",   "log_E_v2.txt",  0, "NONE"),
]

hdr = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="4472C4")
border = Border(*[Side(style="thin", color="D9D9D9")]*4)
def style_header(ws, n):
    for c in range(1, n+1):
        x = ws.cell(1, c); x.font = hdr; x.fill = fill; x.alignment = Alignment(horizontal="center")

wb = Workbook(); ringkas = []
for scen, fn, tgt, exp in LOGS:
    if not os.path.exists(fn):
        print(f"  [lewati] {fn} tidak ada"); continue
    txt = open(fn, encoding="utf-8", errors="ignore").read().splitlines()
    anoms, commits = [], []
    for ln in txt:
        m = RE_ANOM.search(ln)
        if m:
            anoms.append((int(m.group(1)), int(m.group(2)), m.group(3).upper(), (m.group(4) or "").strip())); continue
        m = RE_COMMIT.search(ln)
        if m: commits.append((int(m.group(1)), int(m.group(2))))
    def commit_for(dt, sl):
        c = [ct for ct, cs in commits if cs == sl and ct >= dt]; return min(c) if c else None
    # SHEET: tiap baris anomali (lengkap, bukan rata-rata)
    ws = wb.create_sheet(f"Log_{scen}")
    ws.append(["No","t_deteksi(ms)","slave","jenis","t_catat(ms)","response(ms)","detail"]); style_header(ws, 7)
    n = 0; resp = []
    for dt, sl, jn, det in anoms:
        ct = commit_for(dt, sl); r = (ct - dt) if ct else ""
        n += 1; ws.append([n, dt, sl, jn, ct or "", r, det])
        if sl == tgt and (exp == "NONE" or jn == exp) and isinstance(r, int): resp.append(r)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=7):
        for c in row: c.border = border
    det_count = len([a for a in anoms if a[1] == tgt and (exp == "NONE" or a[2] == exp)])
    ringkas.append([scen, exp, det_count,
                    round(statistics.median(resp),1) if resp else "-",
                    round(statistics.mean(resp),1) if resp else "-",
                    round(statistics.pstdev(resp),1) if len(resp) > 1 else "-",
                    len(resp)])

# SHEET ringkasan di depan
ws0 = wb["Sheet"]; ws0.title = "Ringkasan"
ws0.append(["Skenario","Jenis target","Terdeteksi","Response median(ms)","mean(ms)","SD(ms)","n_response"]); style_header(ws0, 7)
for r in ringkas: ws0.append(r)
for row in ws0.iter_rows(min_row=1, max_row=ws0.max_row, max_col=7):
    for c in row: c.border = border

for ws in wb.worksheets:
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(w, 50)

out = sys.argv[1] if len(sys.argv) > 1 else "hasil_pengujian.xlsx"
wb.save(out)
print(f"Excel dibuat: {out}  ({len(ringkas)} skenario, tiap log = 1 sheet lengkap)")
for r in ringkas: print("  ", r)
