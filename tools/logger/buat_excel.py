#!/usr/bin/env python3
# buat_excel.py — gabung confusion.csv + response_times.csv → satu file Excel
#   berisi: data mentah, ringkasan per skenario, confusion matrix + metrik.
import csv, statistics, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

if not os.path.exists("confusion.csv"):
    print("[ERROR] confusion.csv belum ada. Jalankan parse_serial.py dulu."); raise SystemExit(1)

conf = list(csv.DictReader(open("confusion.csv", encoding="utf-8")))
resp = list(csv.DictReader(open("response_times.csv", encoding="utf-8"))) if os.path.exists("response_times.csv") else []

wb = Workbook()
hdr = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="4472C4")
def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(1, c); cell.font = hdr; cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

# Sheet 1 — Ringkasan per skenario
ws1 = wb.active; ws1.title = "Ringkasan"
ws1.append(["Skenario","Target","Diharapkan","TP","FP","FN","TN","Deteksi","Response mean (ms)","SD (ms)"])
style_header(ws1, 10)
for r in conf:
    ws1.append([r["Skenario"], r["Target"], r["Expected"], int(r["TP"]), int(r["FP"]),
                int(r["FN"]), int(r["TN"]), r["n_resp"], r["mean_ms"], r["sd_ms"]])

# Sheet 2 — Confusion matrix + metrik (total)
ws2 = wb.create_sheet("Metrik")
TP=sum(int(r["TP"]) for r in conf); FP=sum(int(r["FP"]) for r in conf)
FN=sum(int(r["FN"]) for r in conf); TN=sum(int(r["TN"]) for r in conf)
f=lambda n,d: n/d if d else 0.0
DR,FPR,PRE=f(TP,TP+FN),f(FP,FP+TN),f(TP,TP+FP); ACC=f(TP+TN,TP+TN+FP+FN); F1=f(2*PRE*DR,PRE+DR)
ws2.append(["Confusion Matrix",""]); ws2["A1"].font=Font(bold=True)
ws2.append(["TP (benar tangkap)", TP]); ws2.append(["FP (alarm palsu)", FP])
ws2.append(["FN (lolos)", FN]); ws2.append(["TN (benar diam)", TN])
ws2.append([])
ws2.append(["Metrik","Nilai","Persamaan"]); ws2["A7"].font=Font(bold=True)
ws2.append(["Detection Rate", f"{DR*100:.2f}%", "TP/(TP+FN)  (3.1)"])
ws2.append(["False Positive Rate", f"{FPR*100:.2f}%", "FP/(FP+TN)  (3.2)"])
ws2.append(["Precision", f"{PRE*100:.2f}%", "TP/(TP+FP)  (3.3)"])
ws2.append(["Accuracy", f"{ACC*100:.2f}%", "(TP+TN)/total  (3.4)"])
ws2.append(["F1-score", f"{F1*100:.2f}%", "2PR/(P+R)  (3.5)"])
allr=[float(r["response_ms"]) for r in resp if r["response_ms"].strip()]
if allr:
    ws2.append([]); ws2.append(["Response time (ms)","Nilai"]);
    ws2.append(["mean", round(statistics.mean(allr),1)])
    ws2.append(["median", round(statistics.median(allr),1)])
    ws2.append(["SD", round(statistics.pstdev(allr),1) if len(allr)>1 else 0])
    ws2.append(["min", min(allr)]); ws2.append(["max", max(allr)])

# Sheet 3 — Data mentah
ws3 = wb.create_sheet("Data Mentah")
ws3.append(["Skenario","slave","jenis","t_deteksi(ms)","t_catat(ms)","response(ms)"])
style_header(ws3, 6)
for r in resp:
    ws3.append([r["Skenario"], r["slave"], r["jenis"], r["t_deteksi"], r["t_catat"], r["response_ms"]])

for ws in wb.worksheets:
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(w, 40)

out = "hasil_pengujian.xlsx"
wb.save(out)
print(f"Excel dibuat: {out}")
print(f"  Confusion: TP={TP} FP={FP} FN={FN} TN={TN}")
print(f"  DR={DR*100:.1f}% FPR={FPR*100:.1f}% Precision={PRE*100:.1f}% Accuracy={ACC*100:.1f}% F1={F1*100:.1f}%")
